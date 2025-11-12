#models/reports.py
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from datetime import datetime
import os
import time
import sys
import subprocess
import platform

if platform.system() == "Windows":
    try:
        import win32print
        import win32api
    except ImportError:
        win32print = None
        win32api = None
        print("Advertencia: pywin32 no está instalado. La impresión en Windows puede fallar. Instale con: pip install pywin32")

from typing import Dict, List
from modules.models import obtener_recibo_por_id, obtener_configuracion

# ✅ AGREGAR ESTOS IMPORTS PARA LA FUNCIÓN DE EXCEL
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# ==================== CONFIGURACIÓN DE RECIBO ====================

# IMPORTANTE: Recibo en formato 1/3 carta - ORIENTACIÓN VERTICAL
RECIBO_ANCHO = 21.6 * cm
RECIBO_ALTO = 9.1 * cm

# Ruta del logo
LOGO_PATH = os.path.join('assets', 'lagoo.png')

# ==================== UTILIDADES DE IMPRESIÓN (Windows) ====================

def _buscar_sumatra() -> str | None:
    """
    Busca SumatraPDF en rutas comunes (x64/x86) y retorna la ruta si existe.
    """
    posibles = [
        r"C:\Program Files\SumatraPDF\SumatraPDF.exe",
        r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\SumatraPDF\SumatraPDF.exe"),
        os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\WindowsApps\SumatraPDF.exe"),
    ]
    for p in posibles:
        if os.path.exists(p):
            return p
    return None

def _imprimir_pdf_windows(ruta_pdf: str, impresora: str | None = None) -> None:
    """
    Imprime en Windows con múltiples fallbacks:
    1) SumatraPDF (mejor opción)
    2) Abrir PDF y dejar que el usuario imprima manualmente
    """
    import subprocess
    
    # 1) Intentar con SumatraPDF
    sumatra = _buscar_sumatra()
    if sumatra:
        try:
            args = [sumatra]
            if impresora:
                args += ["-print-to", impresora]
            else:
                args += ["-print-to-default"]
            args += ["-exit-on-print", ruta_pdf]
            subprocess.run(args, check=True, timeout=10)
            print(f"✓ Impreso con SumatraPDF: {ruta_pdf}")
            return
        except Exception as e:
            print(f"⚠ SumatraPDF falló: {e}")
    
    # 2) Fallback: Solo ABRIR el PDF (el usuario imprime manualmente)
    try:
        os.startfile(ruta_pdf)
        print(f"⚠ PDF abierto para impresión manual: {ruta_pdf}")
        # Mostrar mensaje al usuario
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "Impresión Manual", 
            "El PDF se ha abierto.\n\n"
            "Por favor, presiona Ctrl+P para imprimir manualmente.\n\n"
            "Recomendación: Instala SumatraPDF para impresión automática."
        )
        root.destroy()
        return
    except Exception as e:
        raise RuntimeError(
            f"No se puede imprimir en Windows.\n\n"
            f"Soluciones:\n"
            f"1. Instala SumatraPDF (https://www.sumatrapdfreader.org/)\n"
            f"2. Configura una impresora predeterminada en Windows\n"
            f"3. Asocia PDFs con Adobe Reader\n\n"
            f"Error: {e}"
        )


# ==================== GENERACIÓN DE RECIBOS ====================

def generar_recibo_pdf(recibo_id: int, es_reimpresion: bool = False) -> str:
    """Genera el PDF de un recibo en formato 1/3 carta - VERTICAL"""
    recibo = obtener_recibo_por_id(recibo_id)
    if not recibo:
        raise ValueError("Recibo no encontrado")

    nombre_oficina = obtener_configuracion('nombre_oficina') or 'ASOCIACIÓN DE RIEGO'
    ubicacion = obtener_configuracion('ubicacion') or 'Tezontepec de Aldama, Hgo.'

    recibos_dir = os.path.join('database', 'recibos')
    os.makedirs(recibos_dir, exist_ok=True)

    sufijo = '_REIMPRESION' if es_reimpresion else ''
    filename = f"recibo_{recibo['folio']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{sufijo}.pdf"
    filepath = os.path.join(recibos_dir, filename)

    # Crear canvas con tamaño en orientación VERTICAL
    c = canvas.Canvas(filepath, pagesize=(RECIBO_ANCHO, RECIBO_ALTO))
    _dibujar_recibo_principal(c, recibo, nombre_oficina, ubicacion, es_reimpresion)
    c.save()

    return filepath

def generar_recibo_pdf_temporal(recibo_id: int, es_reimpresion: bool = False) -> str:
    """
    Genera el PDF de un recibo TEMPORAL que será eliminado después de imprimir.
    No lo guarda permanentemente en database/recibos/
    """
    recibo = obtener_recibo_por_id(recibo_id)
    if not recibo:
        raise ValueError("Recibo no encontrado")

    nombre_oficina = obtener_configuracion('nombre_oficina') or 'ASOCIACIÓN DE RIEGO'
    ubicacion = obtener_configuracion('ubicacion') or 'Tezontepec de Aldama, Hgo.'

    # Crear carpeta temporal en /tmp (Mac/Linux) o %TEMP% (Windows)
    if platform.system() == "Windows":
        temp_dir = os.path.join(os.environ.get('TEMP', os.getcwd()), 'recibos_temp')
    else:
        temp_dir = os.path.join('/tmp', 'recibos_temp')

    os.makedirs(temp_dir, exist_ok=True)

    sufijo = '_REIMPRESION' if es_reimpresion else ''
    filename = f"recibo_{recibo['folio']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{sufijo}.pdf"
    filepath = os.path.join(temp_dir, filename)

    c = canvas.Canvas(filepath, pagesize=(RECIBO_ANCHO, RECIBO_ALTO))
    _dibujar_recibo_principal(c, recibo, nombre_oficina, ubicacion, es_reimpresion)
    c.save()

    return filepath

# ==================== IMPRESIÓN (TEMPORALES) ====================

def imprimir_recibo_y_limpiar(pdf_path: str):
    """
    Intenta imprimir un archivo PDF en Windows o macOS y luego lo elimina.
    """
    sistema = platform.system()
    print(f"Intentando imprimir en {sistema} desde: {pdf_path}")

    if not os.path.exists(pdf_path):
        print(f"Error: El archivo PDF no existe: {pdf_path}")
        return

    impreso = False

    if sistema == "Windows":
        if win32print and win32api:
            try:
                # Usar win32api para imprimir
                win32api.ShellExecute(0, "print", pdf_path, None, ".", 0)
                print(f"Comando de impresión enviado para win32: {pdf_path}")
                impreso = True
            except Exception as e:
                 print(f"Error con win32api.ShellExecute: {e}")
        else:
             print("pywin32 no disponible. Intentando alternativas...")

        if not impreso:
            # Alternativa menos confiable: os.startfile
            try:
                os.startfile(pdf_path, "print") # Esta acción puede no ser silenciosa
                print(f"Comando de impresión enviado para os.startfile: {pdf_path}")
                impreso = True
            except Exception as e:
                 print(f"Error con os.startfile: {e}")

    elif sistema == "Darwin": # macOS
        try:
            # Usar 'lp' para imprimir en macOS (requiere CUPS instalado, que generalmente lo está)
            subprocess.run(["lp", pdf_path], check=True)
            print(f"PDF impreso en macOS usando lp: {pdf_path}")
            impreso = True
        except subprocess.CalledProcessError as e:
            print(f"Error al imprimir con 'lp' en macOS: {e}")
        except FileNotFoundError:
            print("Comando 'lp' no encontrado en macOS. Verifique la instalación de CUPS o la ruta.")

    else: # Linux u Otro
        try:
            # Usar 'lp' para imprimir en Linux (requiere CUPS instalado)
            subprocess.run(["lp", pdf_path], check=True)
            print(f"PDF impreso en Linux usando lp: {pdf_path}")
            impreso = True
        except subprocess.CalledProcessError as e:
            print(f"Error al imprimir con 'lp' en Linux: {e}")
        except FileNotFoundError:
            print("Comando 'lp' no encontrado. Verifique la instalación de CUPS o la ruta.")

    # Esperar un momento para que el comando de impresión se procese
    import time
    time.sleep(1) # Ajusta si es necesario

    # Eliminar el archivo temporal después del intento de impresión
    try:
        os.remove(pdf_path)
        print(f"Archivo temporal eliminado: {pdf_path}")
    except OSError as e:
        print(f"Error al eliminar archivo temporal {pdf_path}: {e}")

# ==================== DIBUJO DE RECIBO ====================

def _dibujar_recibo_principal(c, recibo: Dict, nombre_oficina: str, ubicacion: str, es_reimpresion: bool):
    """
    Dibuja el recibo principal - MARCA DE AGUA AL FINAL
    """
    
    # ===== COLORES =====
    COLOR_VERDE = colors.HexColor('#B8D1BF')
    COLOR_BEIGE = colors.HexColor('#FFFFFF')
    COLOR_BEIGE_OSCURO = colors.HexColor('#C9B99A')
    COLOR_TEXTO = colors.HexColor('#2C3E2E')
    COLOR_TEXTO_GRIS = colors.HexColor('#666666')
    
    # ===== FONDO BEIGE =====
    c.setFillColor(COLOR_BEIGE)
    c.roundRect(0.15*cm, 0.15*cm, RECIBO_ANCHO - 0.3*cm, RECIBO_ALTO - 0.3*cm, 
                0.5*cm, stroke=0, fill=1)
    
    # ===== HEADER VERDE (más alto) =====
    c.setFillColor(COLOR_VERDE)
    c.roundRect(0.4*cm, RECIBO_ALTO - 2.3*cm, RECIBO_ANCHO - 0.8*cm, 1.9*cm, 
                0.4*cm, stroke=0, fill=1)
    
    # ===== LOGO (más grande) =====
    if os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, 0.7*cm, RECIBO_ALTO - 2.1*cm, 
                       width=1.7*cm, height=1.7*cm, mask='auto')
        except:
            pass
    
    # ===== TÍTULO (texto más grande) =====
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(RECIBO_ANCHO/2 + 0.5*cm, RECIBO_ALTO - 1*cm, 
                       "ASOCIACIÓN DE CAMPESINOS DE BOMBEO Y REBOMBEO")
    
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(RECIBO_ANCHO/2 + 0.5*cm, RECIBO_ALTO - 1.35*cm, 
                       "DEL CERRO DEL XICUCO A.C. M7-1")
    
    c.setFont("Helvetica", 8)
    c.drawCentredString(RECIBO_ANCHO/2 + 0.5*cm, RECIBO_ALTO - 1.75*cm, 
                       "RFC: ACB030619G68")
    
    # ===== SEPARADOR =====
    y_pos = RECIBO_ALTO - 2.45*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.setLineWidth(0.5)
    c.line(0.7*cm, y_pos, RECIBO_ANCHO - 0.7*cm, y_pos)
    
    # ===== CAJA DE DATOS (más alta) =====
    y_pos -= 0.2*cm
    c.setFillColor(colors.white)
    c.roundRect(0.7*cm, y_pos - 1.5*cm, RECIBO_ANCHO - 1.4*cm, 1.4*cm, 
                0.25*cm, stroke=1, fill=1)
    
    # ===== GRID DE DATOS (texto más grande) =====
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 8.5)
    
    col1 = 1*cm
    col2 = 6.5*cm
    col3 = 12*cm
    col4 = 17.5*cm
    
    row_y = y_pos - 0.45*cm
    
    # FILA 1
    c.drawString(col1, row_y, f"NO. RECIBO: {recibo['folio']}")
    c.drawString(col2, row_y, f"No. Lote: {recibo['numero_lote']}")
    c.drawString(col3, row_y, f"No. Riego: {recibo['numero_riego']}")
    c.drawString(col4, row_y, f"Barrio: {recibo['barrio']}")
    
    row_y -= 0.45*cm
    
    # FILA 2
    col1_fila2 = 1*cm
    col2_fila2 = 8.5*cm
    col3_fila2 = 16*cm
    
    c.drawString(col1_fila2, row_y, f"Cultivo: {recibo['cultivo']}")
    c.drawString(col2_fila2, row_y, f"Superficie: {recibo['superficie']} ha")
    c.drawString(col3_fila2, row_y, f"Ciclo: {recibo['ciclo']}")
    
    # ===== RECIBÍ DE (más espacio y texto más grande) =====
    y_pos = row_y - 0.95*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 8.5)
    c.drawString(0.8*cm, y_pos, "Recibí de:")
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2.2*cm, y_pos, recibo['nombre'].upper())
    
    # ===== CONCEPTO (más espacio) =====
    y_pos -= 0.5*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.line(0.8*cm, y_pos, RECIBO_ANCHO - 0.8*cm, y_pos)
    
    y_pos -= 0.3*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 8)
    c.drawString(0.8*cm, y_pos, "Concepto: Pago de cuota de riego para el ciclo agrícola")
    
    # ===== TOTAL + MONTO (más espacio y texto más grande) =====
    y_pos -= 0.5*cm
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(0.8*cm, y_pos, "TOTAL")
    
    # Caja del monto (más grande)
    monto_x = RECIBO_ANCHO - 5*cm
    c.setFillColor(colors.white)
    c.setStrokeColor(COLOR_VERDE)
    c.setLineWidth(1.5)
    c.roundRect(monto_x, y_pos - 0.35*cm, 4.2*cm, 0.75*cm, 
                0.25*cm, stroke=1, fill=1)
    
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 6.5)
    c.drawString(monto_x + 0.2*cm, y_pos + 0.15*cm, "(pago en efectivo)")
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 15)
    c.drawRightString(monto_x + 4*cm, y_pos - 0.15*cm, f"${recibo['costo']:.2f}")
    
    # ===== FOOTER (más espacio) =====
    y_pos -= 0.9*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.setLineWidth(0.5)
    c.line(0.8*cm, y_pos, RECIBO_ANCHO - 0.8*cm, y_pos)
    
    y_pos -= 0.3*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 7.5)
    
    fecha_obj = datetime.strptime(recibo['fecha'], '%Y-%m-%d')
    c.drawString(0.8*cm, y_pos, 
                f"C. Juan Aldama #25, Col. Centro, Tezontepec de Aldama. Fecha: {fecha_obj.strftime('%d/%m/%Y')}")
    
    y_pos -= 0.28*cm
    hora_obj = datetime.strptime(recibo['hora'], '%H:%M:%S')
    am_pm = "p.m." if hora_obj.hour >= 12 else "a.m."
    hora_12 = hora_obj.hour if hora_obj.hour <= 12 else hora_obj.hour - 12
    if hora_12 == 0:
        hora_12 = 12
    
    c.drawString(0.8*cm, y_pos, f"Hora: {hora_12:02d}:{hora_obj.minute:02d}:{hora_obj.second:02d} {am_pm}")
    
    # Firma (más espacio)
    c.drawRightString(RECIBO_ANCHO - 0.8*cm, y_pos + 0.28*cm, "Firma Recaudador")
    c.line(RECIBO_ANCHO - 4*cm, y_pos + 0.18*cm, RECIBO_ANCHO - 0.8*cm, y_pos + 0.18*cm)
    
    # ===== LEYENDA LEGAL (texto más grande y más espacio) =====
    y_pos -= 0.45*cm
    c.setFont("Helvetica", 6)
    
    c.drawString(0.7*cm, y_pos, 
                "Este recibo ampara el pago de cuota ordinaria destinada exclusivamente al mantenimiento y operación del módulo de riego, conforme al régimen fiscal")
    y_pos -= 0.22*cm
    c.drawString(0.7*cm, y_pos,
                "de personas morales con fines no lucrativos. Exento de IVA y de ISR conforme a los artículos 79 y 80 de la Ley del ISR y al artículo 15, fracción XII de la Ley del IVA.")
    
    # ===== MARCA DE AGUA (CENTRADA VERTICAL Y HORIZONTALMENTE) =====
    if es_reimpresion:
        c.saveState()
        c.setFont("Helvetica-Bold", 32)
        c.setFillColorRGB(0.9, 0.9, 0.9)
        c.rotate(30)
        c.drawString(8 * cm, 0.1 * cm, "REIMPRESIÓN") # Cambiado de 0.5*cm a 0.1*cm
        c.restoreState()

# ==================== REPORTE DIARIO ====================

def generar_reporte_diario(fecha: str, recibos: List[Dict]) -> str:
    """Genera un reporte PDF del día con todos los recibos - ORIENTACIÓN VERTICAL"""
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    filename = f"reporte_diario_{fecha.replace('-', '')}.pdf"
    filepath = os.path.join(reportes_dir, filename)

    # Canvas con tamaño PORTRAIT (vertical) - letter = (8.5 x 11 inches)
    c = canvas.Canvas(filepath, pagesize=letter)

    nombre_oficina = obtener_configuracion('nombre_oficina') or 'ASOCIACIÓN DE RIEGO'

    # --- Añadir Logo al Reporte Diario ---
    if os.path.exists(LOGO_PATH):
        try:
            logo_width = 2 * cm
            logo_height = 2 * cm
            c.drawImage(LOGO_PATH, 2*cm, letter[1] - 3*cm, width=logo_width, height=logo_height, mask='auto')
        except Exception as e:
            print(f"Error al añadir logo al reporte diario: {e}")
    # --------------------------------------

    y_pos = letter[1] - 2*cm
    margen_izq = 2*cm
    margen_der = letter[0] - 2*cm

    # ENCABEZADO
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(letter[0]/2, y_pos, nombre_oficina.upper())
    y_pos -= 0.7*cm

    c.setFont("Helvetica-Bold", 12)
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
    c.drawCentredString(letter[0]/2, y_pos, f"REPORTE DIARIO - {fecha_obj.strftime('%d/%m/%Y')}")
    y_pos -= 0.7*cm

    c.setFont("Helvetica", 10)
    c.drawCentredString(letter[0]/2, y_pos, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    y_pos -= 1*cm

    # TABLA DE RECIBOS
    if not recibos:
        c.setFont("Helvetica", 11)
        c.drawCentredString(letter[0]/2, y_pos, "No hay recibos registrados en este día")
    else:
        c.setFont("Helvetica-Bold", 8)
        col_widths = [1.5*cm, 2*cm, 1.5*cm, 5*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2*cm]
        col_x = [margen_izq]
        for w in col_widths[:-1]:
            col_x.append(col_x[-1] + w)

        headers = ["Folio", "Hora", "Lote", "Nombre", "Cultivo", "Riego", "Acción", "Monto"]
        for i, header in enumerate(headers):
            c.drawString(col_x[i], y_pos, header)
        y_pos -= 0.3*cm
        c.line(margen_izq, y_pos, margen_der, y_pos)
        y_pos -= 0.4*cm

        c.setFont("Helvetica", 7)
        total_dia = 0

        for recibo in recibos:
            if y_pos < 3*cm:
                c.showPage()
                # Volver a dibujar el logo en la nueva página si es necesario
                if os.path.exists(LOGO_PATH):
                    try:
                        logo_width = 2 * cm
                        logo_height = 2 * cm
                        c.drawImage(LOGO_PATH, 2*cm, letter[1] - 3*cm, width=logo_width, height=logo_height, mask='auto')
                    except Exception as e:
                        print(f"Error al añadir logo a nueva página del reporte: {e}")

                y_pos = letter[1] - 2*cm
                c.setFont("Helvetica-Bold", 8)
                for i, header in enumerate(headers):
                    c.drawString(col_x[i], y_pos, header)
                y_pos -= 0.3*cm
                c.line(margen_izq, y_pos, margen_der, y_pos)
                y_pos -= 0.4*cm
                c.setFont("Helvetica", 7)

            c.drawString(col_x[0], y_pos, str(recibo['folio']))
            c.drawString(col_x[1], y_pos, recibo['hora'][:5])
            c.drawString(col_x[2], y_pos, recibo['numero_lote'])
            nombre = recibo['nombre'][:30] if len(recibo['nombre']) > 30 else recibo['nombre']
            c.drawString(col_x[3], y_pos, nombre)
            c.drawString(col_x[4], y_pos, recibo['cultivo'])
            c.drawString(col_x[5], y_pos, str(recibo['numero_riego']))
            tipo = "Nueva" if recibo['tipo_accion'] == 'Nueva siembra' else "Adicional"
            c.drawString(col_x[6], y_pos, tipo)
            c.drawRightString(col_x[7] + 2*cm, y_pos, f"${recibo['costo']:.2f}")
            total_dia += recibo['costo']
            y_pos -= 0.35*cm

        y_pos -= 0.2*cm
        c.line(margen_izq, y_pos, margen_der, y_pos)
        y_pos -= 0.5*cm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(margen_izq, y_pos, f"TOTAL DEL DÍA:")
        c.drawRightString(margen_der, y_pos, f"${total_dia:.2f}")
        y_pos -= 0.5*cm

        c.setFont("Helvetica", 9)
        c.drawString(margen_izq, y_pos, f"Total de recibos emitidos: {len(recibos)}")

    c.save()
    return filepath

# ==================== IMPRESIÓN DIRECTA (NO TEMPORALES) ====================

def imprimir_recibo(ruta_pdf: str, impresora: str = None):
    """Envía el PDF a la impresora sin eliminar el archivo (para PDFs no temporales)"""
    if not os.path.exists(ruta_pdf):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_pdf}")

    sistema = platform.system()

    try:
        if sistema == 'Windows':
            _imprimir_pdf_windows(ruta_pdf, impresora)
        elif sistema == 'Darwin':
            cmd = ['lp', ruta_pdf] if not impresora else ['lp', '-d', impresora, ruta_pdf]
            subprocess.run(cmd, check=True)
        else:
            cmd = ['lp', ruta_pdf] if not impresora else ['lp', '-d', impresora, ruta_pdf]
            subprocess.run(cmd, check=True)
        return True
    except Exception as e:
        print(f"Error al imprimir: {e}")
        return False

# ==================== LISTA DE IMPRESORAS ====================

def obtener_impresoras_disponibles() -> List[str]:
    """
    Lista impresoras disponibles:
    - Windows: pywin32 si está, si no PowerShell/WMIC (fallback).
    - macOS/Linux: lpstat.
    """
    try:
        sistema = platform.system()

        if sistema == "Windows":
            # 1) Intento con pywin32
            try:
                import win32print  # type: ignore
                impresoras = win32print.EnumPrinters(2)
                return [imp[2] for imp in impresoras if len(imp) >= 3]
            except Exception:
                pass  # seguir al fallback

            # 2) PowerShell (Get-CimInstance)
            try:
                ps = [
                    "powershell", "-NoProfile", "-Command",
                    "Get-CimInstance -ClassName Win32_Printer | Select-Object -ExpandProperty Name"
                ]
                r = subprocess.run(ps, capture_output=True, text=True, timeout=5)
                names = [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
                if names:
                    return names
            except Exception:
                pass

            # 3) WMIC (legacy pero aún frecuente)
            try:
                r = subprocess.run(["wmic", "printer", "get", "name"],
                                   capture_output=True, text=True, timeout=5)
                names = [ln.strip() for ln in r.stdout.splitlines()[1:] if ln.strip()]
                if names:
                    return names
            except Exception:
                pass

            return ["Impresora por defecto"]

        elif sistema == "Darwin":  # macOS
            r = subprocess.run(["lpstat", "-p", "-d"], capture_output=True, text=True, timeout=5)
            impresoras = []
            for ln in r.stdout.splitlines():
                if ln.startswith("printer"):
                    partes = ln.split()
                    if len(partes) >= 2:
                        impresoras.append(partes[1])
            return impresoras if impresoras else ["Impresora por defecto"]

        else:  # Linux
            r = subprocess.run(["lpstat", "-p", "-d"], capture_output=True, text=True, timeout=5)
            impresoras = []
            for ln in r.stdout.splitlines():
                if ln.startswith("printer"):
                    partes = ln.split()
                    if len(partes) >= 2:
                        impresoras.append(partes[1])
            return impresoras if impresoras else ["Impresora por defecto"]

    except Exception:
        return ["Impresora por defecto"]

# ==================== ABRIR PDF ====================

def abrir_pdf(ruta_pdf: str):
    """Abre el PDF con el visor predeterminado del sistema"""
    if not os.path.exists(ruta_pdf):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_pdf}")

    sistema = platform.system()

    try:
        if sistema == 'Windows':
            os.startfile(ruta_pdf)  # depende de asociación, solo para ver
        elif sistema == 'Darwin':
            subprocess.run(['open', ruta_pdf])
        else:
            subprocess.run(['xdg-open', ruta_pdf])
        return True
    except Exception as e:
        print(f"Error al abrir PDF: {e}")
        return False

# ==================== EXPORTACIÓN A EXCEL ====================

def exportar_a_excel(recibos: List[Dict], filename: str) -> str:
    """Exporta una lista de recibos a un archivo Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        # Crear libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recibos"

        # Encabezados
        headers = ['Folio', 'Fecha', 'Hora', 'Lote', 'Nombre', 'Localidad', 'Barrio',
                   'Superficie', 'Cultivo', 'Riego No.', 'Acción', 'Costo']

        # Estilo de encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row, recibo in enumerate(recibos, 2):
            ws.cell(row=row, column=1, value=recibo['folio'])
            ws.cell(row=row, column=2, value=recibo['fecha'])
            ws.cell(row=row, column=3, value=recibo['hora'])
            ws.cell(row=row, column=4, value=recibo['numero_lote'])
            ws.cell(row=row, column=5, value=recibo['nombre'])
            ws.cell(row=row, column=6, value=recibo['localidad'])
            ws.cell(row=row, column=7, value=recibo['barrio'])
            ws.cell(row=row, column=8, value=recibo['superficie'])
            ws.cell(row=row, column=9, value=recibo['cultivo'])
            ws.cell(row=row, column=10, value=recibo['numero_riego'])
            ws.cell(row=row, column=11, value=recibo['tipo_accion'])
            ws.cell(row=row, column=12, value=recibo['costo'])

        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column].width = adjusted_width

        # Guardar archivo
        reportes_dir = os.path.join('database', 'reportes')
        os.makedirs(reportes_dir, exist_ok=True)
        filepath = os.path.join(reportes_dir, filename)
        wb.save(filepath)

        return filepath

    except ImportError:
        raise ImportError("La librería 'openpyxl' no está instalada. Instálala con: pip install openpyxl")
    except Exception as e:
        raise Exception(f"Error al exportar a Excel: {e}")
    
def generar_corte_caja_excel(fecha: str, recibos: List[Dict]) -> str:
    """
    Genera un archivo Excel con el corte de caja del día.
    
    Args:
        fecha: Fecha en formato YYYY-MM-DD
        recibos: Lista de recibos del día
    
    Returns:
        Ruta del archivo Excel generado
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Crear directorio si no existe
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    # Nombre del archivo
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
    fecha_str = fecha_obj.strftime('%Y%m%d')
    nombre_archivo = f"corte_caja_{fecha_str}.xlsx"
    ruta_excel = os.path.join(reportes_dir, nombre_archivo)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corte de Caja"
    
    # ===== ESTILOS =====
    titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    total_font = Font(name='Calibri', size=12, bold=True)
    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== ENCABEZADO =====
    ws.merge_cells('A1:H1')
    cell_titulo = ws['A1']
    cell_titulo.value = 'CORTE DE CAJA'
    cell_titulo.font = titulo_font
    cell_titulo.fill = titulo_fill
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:H2')
    cell_fecha = ws['A2']
    cell_fecha.value = f"Fecha: {fecha_obj.strftime('%d/%m/%Y')}"
    cell_fecha.font = Font(name='Calibri', size=12, bold=True)
    cell_fecha.alignment = Alignment(horizontal='center')
    
    nombre_oficina = obtener_configuracion('nombre_oficina') or 'SISTEMA DE RIEGO'
    ws.merge_cells('A3:H3')
    cell_oficina = ws['A3']
    cell_oficina.value = nombre_oficina
    cell_oficina.font = Font(name='Calibri', size=11, italic=True)
    cell_oficina.alignment = Alignment(horizontal='center')
    
    # ===== CABECERAS =====
    headers = ['Folio', 'Lote', 'Nombre', 'Cultivo', 'Superficie', 'Riego', 'Monto', 'Hora']
    row_num = 5
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # ===== DATOS =====
    row_num = 6
    total_monto = 0
    
    for recibo in recibos:
        if recibo.get('eliminado'):
            continue
            
        ws.cell(row=row_num, column=1, value=recibo['folio']).border = border
        ws.cell(row=row_num, column=2, value=recibo['numero_lote']).border = border
        ws.cell(row=row_num, column=3, value=recibo['nombre']).border = border
        ws.cell(row=row_num, column=4, value=recibo['cultivo']).border = border
        
        sup_cell = ws.cell(row=row_num, column=5, value=recibo['superficie'])
        sup_cell.border = border
        sup_cell.alignment = Alignment(horizontal='right')
        
        ws.cell(row=row_num, column=6, value=recibo['numero_riego']).border = border
        
        monto_cell = ws.cell(row=row_num, column=7, value=recibo['costo'])
        monto_cell.border = border
        monto_cell.number_format = '$#,##0.00'
        monto_cell.alignment = Alignment(horizontal='right')
        
        ws.cell(row=row_num, column=8, value=recibo['hora']).border = border
        
        total_monto += recibo['costo']
        row_num += 1
    
    # ===== TOTALES =====
    row_num += 1
    ws.merge_cells(f'A{row_num}:F{row_num}')
    cell_total_label = ws.cell(row=row_num, column=1)
    cell_total_label.value = 'TOTAL DEL DÍA:'
    cell_total_label.font = total_font
    cell_total_label.fill = total_fill
    cell_total_label.alignment = Alignment(horizontal='right')
    cell_total_label.border = border
    
    cell_total_monto = ws.cell(row=row_num, column=7)
    cell_total_monto.value = total_monto
    cell_total_monto.font = total_font
    cell_total_monto.fill = total_fill
    cell_total_monto.number_format = '$#,##0.00'
    cell_total_monto.alignment = Alignment(horizontal='right')
    cell_total_monto.border = border
    
    ws.cell(row=row_num, column=8).border = border
    
    # ===== ESTADÍSTICAS =====
    row_num += 2
    ws.cell(row=row_num, column=1, value='ESTADÍSTICAS:').font = Font(bold=True)
    row_num += 1
    
    ws.cell(row=row_num, column=1, value=f"Total de recibos:")
    ws.cell(row=row_num, column=2, value=len([r for r in recibos if not r.get('eliminado')]))
    row_num += 1
    
    # Recibos por tipo
    nuevas_siembras = len([r for r in recibos if r['tipo_accion'] == 'Nueva siembra' and not r.get('eliminado')])
    riegos_adicionales = len([r for r in recibos if r['tipo_accion'] == 'Riego adicional' and not r.get('eliminado')])
    
    ws.cell(row=row_num, column=1, value=f"Nuevas siembras:")
    ws.cell(row=row_num, column=2, value=nuevas_siembras)
    row_num += 1
    
    ws.cell(row=row_num, column=1, value=f"Riegos adicionales:")
    ws.cell(row=row_num, column=2, value=riegos_adicionales)
    row_num += 1
    
    # ===== PIE DE PÁGINA =====
    row_num += 2
    ws.merge_cells(f'A{row_num}:H{row_num}')
    cell_generado = ws.cell(row=row_num, column=1)
    cell_generado.value = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}"
    cell_generado.font = Font(name='Calibri', size=9, italic=True, color='808080')
    cell_generado.alignment = Alignment(horizontal='center')
    
    # ===== AJUSTAR ANCHOS DE COLUMNA =====
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    # Guardar archivo
    wb.save(ruta_excel)
    
    print(f"✅ Corte de caja Excel generado: {ruta_excel}")
    return ruta_excel

def generar_pdf_estadisticas(estadisticas: Dict, estadisticas_cultivo: List[Dict]) -> str:
    """
    Genera un PDF profesional con las estadísticas del sistema.
    
    Args:
        estadisticas: Diccionario con estadísticas generales
        estadisticas_cultivo: Lista de estadísticas por cultivo
    
    Returns:
        Ruta del archivo PDF generado
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import Table, TableStyle
    
    # Crear directorio si no existe
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    # Nombre del archivo
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"estadisticas_{fecha_str}.pdf"
    ruta_pdf = os.path.join(reportes_dir, nombre_archivo)
    
    # Crear PDF
    c = pdf_canvas.Canvas(ruta_pdf, pagesize=letter)
    ancho, alto = letter
    
    # ===== ENCABEZADO =====
    y_pos = alto - 2*cm
    
    # Logo si existe
    if os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, 2*cm, y_pos - 1.5*cm, width=2*cm, height=2*cm, mask='auto')
        except:
            pass
    
    # Título
    c.setFont("Helvetica-Bold", 18)
    c.drawString(5*cm, y_pos, "ESTADÍSTICAS DEL SISTEMA")
    y_pos -= 0.5*cm
    
    nombre_oficina = obtener_configuracion('nombre_oficina') or 'SISTEMA DE RIEGO'
    c.setFont("Helvetica", 12)
    c.drawString(5*cm, y_pos, nombre_oficina)
    y_pos -= 0.4*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(5*cm, y_pos, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    y_pos -= 1*cm
    
    # Línea separadora
    c.line(2*cm, y_pos, ancho - 2*cm, y_pos)
    y_pos -= 1*cm
    
    # ===== ESTADÍSTICAS GENERALES =====
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, y_pos, "📊 ESTADÍSTICAS GENERALES")
    y_pos -= 0.7*cm
    
    # Crear tabla de estadísticas generales
    datos_generales = [
        ['Indicador', 'Valor'],
        ['Total de Campesinos', str(estadisticas.get('total_campesinos', 0))],
        ['Total de Lotes', str(estadisticas.get('total_lotes', 0))],
        ['Superficie Total', f"{estadisticas.get('superficie_total', 0):.2f} ha"],
        ['Siembras Activas', str(estadisticas.get('siembras_activas', 0))],
        ['Total de Recibos', str(estadisticas.get('total_recibos', 0))],
        ['Ingresos Totales', f"${estadisticas.get('ingresos_totales', 0):,.2f}"],
    ]
    
    tabla_general = Table(datos_generales, colWidths=[8*cm, 6*cm])
    tabla_general.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    
    tabla_general.wrapOn(c, ancho, alto)
    tabla_general.drawOn(c, 2*cm, y_pos - 4*cm)
    y_pos -= 5*cm
    
    # ===== ESTADÍSTICAS POR CULTIVO =====
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, y_pos, "🌱 ESTADÍSTICAS POR CULTIVO")
    y_pos -= 0.7*cm
    
    if estadisticas_cultivo:
        datos_cultivos = [['Cultivo', 'Siembras', 'Superficie (ha)', 'Recibos', 'Ingresos']]
        
        for cultivo in estadisticas_cultivo:
            datos_cultivos.append([
                cultivo['cultivo'],
                str(cultivo['num_siembras']),
                f"{cultivo['superficie_total']:.2f}",
                str(cultivo['num_recibos']),
                f"${cultivo['ingresos_totales']:,.2f}"
            ])
        
        tabla_cultivos = Table(datos_cultivos, colWidths=[3.5*cm, 2.5*cm, 3*cm, 2.5*cm, 3.5*cm])
        tabla_cultivos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F5E9')]),
        ]))
        
        altura_tabla = len(datos_cultivos) * 0.6 * cm
        tabla_cultivos.wrapOn(c, ancho, alto)
        tabla_cultivos.drawOn(c, 2*cm, y_pos - altura_tabla)
        y_pos -= altura_tabla + 1*cm
    
    # ===== PIE DE PÁGINA =====
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.grey)
    c.drawCentredString(ancho/2, 1.5*cm, f"Página 1 - Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}")
    
    c.save()
    
    print(f"✅ PDF de estadísticas generado: {ruta_pdf}")
    return ruta_pdf


def generar_pdf_auditoria(registros_auditoria: List[Dict], fecha_inicio=None, fecha_fin=None) -> str:
    """
    Genera un PDF profesional con el historial de auditoría.
    
    Args:
        registros_auditoria: Lista de registros de auditoría
        fecha_inicio: Fecha de inicio del rango (opcional)
        fecha_fin: Fecha de fin del rango (opcional)
    
    Returns:
        Ruta del archivo PDF generado
    """
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import Table, TableStyle, PageBreak
    
    # Crear directorio si no existe
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    # Nombre del archivo
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"auditoria_{fecha_str}.pdf"
    ruta_pdf = os.path.join(reportes_dir, nombre_archivo)
    
    # Usar orientación horizontal para más espacio
    c = pdf_canvas.Canvas(ruta_pdf, pagesize=landscape(letter))
    ancho, alto = landscape(letter)
    
    # ===== FUNCIÓN PARA DIBUJAR ENCABEZADO =====
    def dibujar_encabezado(c, y_pos):
        # Logo si existe
        if os.path.exists(LOGO_PATH):
            try:
                c.drawImage(LOGO_PATH, 2*cm, y_pos - 1.5*cm, width=1.5*cm, height=1.5*cm, mask='auto')
            except:
                pass
        
        # Título
        c.setFont("Helvetica-Bold", 16)
        c.drawString(4.5*cm, y_pos, "REGISTRO DE AUDITORÍA")
        y_pos -= 0.5*cm
        
        nombre_oficina = obtener_configuracion('nombre_oficina') or 'SISTEMA DE RIEGO'
        c.setFont("Helvetica", 10)
        c.drawString(4.5*cm, y_pos, nombre_oficina)
        y_pos -= 0.3*cm
        
        if fecha_inicio and fecha_fin:
            c.drawString(4.5*cm, y_pos, f"Período: {fecha_inicio} - {fecha_fin}")
            y_pos -= 0.3*cm
        
        c.setFont("Helvetica", 9)
        c.drawString(4.5*cm, y_pos, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        y_pos -= 0.5*cm
        
        c.line(2*cm, y_pos, ancho - 2*cm, y_pos)
        y_pos -= 0.5*cm
        
        return y_pos
    
    # Primera página
    pagina = 1
    y_pos = alto - 1.5*cm
    y_pos = dibujar_encabezado(c, y_pos)
    
    # ===== RESUMEN =====
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, f"📋 RESUMEN: {len(registros_auditoria)} registros totales")
    y_pos -= 1*cm
    
    # ===== TABLA DE REGISTROS =====
    registros_por_pagina = 20
    
    for i in range(0, len(registros_auditoria), registros_por_pagina):
        if i > 0:
            # Nueva página
            c.showPage()
            pagina += 1
            y_pos = alto - 1.5*cm
            y_pos = dibujar_encabezado(c, y_pos)
        
        chunk = registros_auditoria[i:i+registros_por_pagina]
        
        datos_tabla = [['Fecha/Hora', 'Tipo', 'Descripción', 'Usuario/Lote']]
        
        for reg in chunk:
            fecha_hora = f"{reg['fecha']} {reg['hora']}"
            tipo = reg['tipo_accion']
            descripcion = reg['descripcion'][:50] + '...' if len(reg['descripcion']) > 50 else reg['descripcion']
            usuario = f"ID: {reg['campesino_id']}" if reg['campesino_id'] else '-'
            
            datos_tabla.append([fecha_hora, tipo, descripcion, usuario])
        
        tabla = Table(datos_tabla, colWidths=[4.5*cm, 4*cm, 10*cm, 3*cm])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F497D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        altura_tabla = (len(datos_tabla) + 1) * 0.5 * cm
        tabla.wrapOn(c, ancho, alto)
        tabla.drawOn(c, 2*cm, y_pos - altura_tabla)
        
        # Pie de página
        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.grey)
        c.drawCentredString(ancho/2, 1*cm, f"Página {pagina} - Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}")
        c.setFillColor(colors.black)
    
    c.save()
    
    print(f"✅ PDF de auditoría generado: {ruta_pdf}")
    return ruta_pdf
