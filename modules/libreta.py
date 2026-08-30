# modules/libreta.py
"""
Generación de la libreta oficial de CONAGUA (FORMA B NUM. 1) en Excel.

Produce la libreta *en blanco*: sólo el padrón (lote, nombre, toma y superficie)
ya transcrito, con las casillas de siembra y de riegos vacías para llenarlas a
mano.

Todo el formato (anchos, alturas, márgenes, escala, saltos de página) está
copiado de 'cambios libreta.xlsx', la plantilla de CONAGUA con los ajustes de
impresión ya resueltos. El logo se extrajo de 'libreta formato.xlsx' y vive en
assets/conagua.jpg.
"""

import os
import re
import shutil
import zipfile
from typing import Dict, List, Optional

from modules.utils import resource_path


def _letra(idx: int) -> str:
    """Letra de columna de Excel a partir de su índice (1 = A)."""
    letra = ''
    while idx > 0:
        idx, resto = divmod(idx - 1, 26)
        letra = chr(ord('A') + resto) + letra
    return letra

# ==================== GEOMETRÍA DE LA LIBRETA ====================

FILAS_POR_BLOQUE = 46          # cada bloque de 46 filas es una página
PERSONAS_POR_HOJA = 10
RENGLONES_POR_PERSONA = 4
FILAS_DATOS = PERSONAS_POR_HOJA * RENGLONES_POR_PERSONA   # 40
PRIMERA_FILA_DATOS = 6         # desplazamiento dentro del bloque

ULTIMA_COLUMNA_CARA_A = 18     # R: el salto de columna va justo detrás
COLUMNA_NUMERO_CARA_B = 19     # S: el consecutivo repetido del reverso
PRIMER_RIEGO_CARA_B = 20       # T: FECHA del 2º riego
TOTAL_COLUMNAS = 37            # hasta la AK, RECIBO del 10º riego

# Anchos de columna, en caracteres. Todos son representables en 1/256 de
# carácter, que es como Excel los guarda.
ANCHOS = {
    'A': 5.6640625,  'B': 6.0,        'C': 25.1640625, 'D': 16.6640625,
    'E': 5.5,        'F': 6.5,        'G': 9.5,        'H': 7.5,
    'I': 6.0,        'J': 5.0,        'K': 4.5,        'L': 7.0,
    'M': 7.6640625,  'N': 7.83203125, 'O': 10.5,       'P': 9.0,
    'Q': 10.6640625, 'R': 13.0,       'S': 6.0,
}
# De la T a la AK, alternando para cada riego 2º a 10º: FECHA y RECIBO.
for _n in range(9):
    ANCHOS[_letra(PRIMER_RIEGO_CARA_B + _n * 2)] = 8.5        # FECHA
    ANCHOS[_letra(PRIMER_RIEGO_CARA_B + _n * 2 + 1)] = 9.0    # RECIBO

# La cara A (A–R) es lo que entra en una hoja al imprimir. Si esta suma crece,
# el área imprimible se queda corta y Excel parte cada página por la mitad.
SUMA_ANCHOS_CARA_A = 163.65234375

# Alturas de fila por desplazamiento dentro del bloque.
ALTURAS = {0: 14.0, 1: 6.75, 2: 14.0, 3: 9.75, 4: 12.75, 5: 11.25}
ALTURA_DATOS = 12.75

# El logo, a 3.43 x 0.46 pulgadas (329 x 44 px a 96 ppp).
LOGO_ANCHO_PX = 329
LOGO_ALTO_PX = 44
LOGO_RUTA = os.path.join('assets', 'conagua.jpg')

# ==================== TEXTOS DE LA PLANTILLA ====================

TITULO = 'DISTRITO DE RIEGO 0003- TULA, 100 ALFAJAYUCAN Y 112 AJACUBA'
# Los dos espacios antes de 2026 están en la plantilla original.
SUBTITULO = ('FORMA B. NUM. 1 EVALUACION DE LA TECNICA AGRICOLA EN LA SECCION '
             'DE RIEGO NUM. 14 CICLO AGRICOLA 2025 -  2026')

# Cabeceras de la primera fila (+4).
CABECERAS_SUP = {
    'A': 'No.', 'B': 'LOTE', 'C': 'NOMBRE DEL USUARIO', 'D': 'TOMA',
    'E': 'TEN.', 'F': 'SUP.', 'G': 'CULTIVO', 'H': 'SUB CICLO',
    'I': 'USO DE SEMILLA', 'L': 'MPIO.', 'M': 'USO DE MAQUINARIA',
    'Q': '1ER. RIEGO', 'S': 'No.',
}
_RIEGOS = ('2DO. RIEGO', '3ER. RIEGO', '4TO. RIEGO', '5T0. RIEGO',
           '6TO. RIEGO', '7MO. RIEGO', '8VO. RIEGO', '9NO. RIEGO',
           '10MO. RIEGO')
for _n, _riego in enumerate(_RIEGOS):
    CABECERAS_SUP[_letra(PRIMER_RIEGO_CARA_B + _n * 2)] = _riego

# Cabeceras de la segunda fila (+5).
CABECERAS_INF = {
    'I': 'CRIO.', 'J': 'MEJ.', 'K': 'VAR.',
    'M': 'PREP.', 'N': 'SIEM.', 'O': 'LAB. CULT.', 'P': 'COS.',
    'Q': 'FECHA', 'R': 'RECIBO',
}
for _n in range(9):
    CABECERAS_INF[_letra(PRIMER_RIEGO_CARA_B + _n * 2)] = 'FECHA'
    CABECERAS_INF[_letra(PRIMER_RIEGO_CARA_B + _n * 2 + 1)] = 'RECIBO'

# Columnas que se combinan verticalmente entre las filas +4 y +5.
COLUMNAS_MERGE_VERTICAL = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'L', 'S')


def verificar_anchos() -> float:
    """
    Comprueba las invariantes de los anchos de columna. Existe para que nadie
    ensanche una columna sin compensar otra: si la suma de la cara A crece, cada
    página impresa se parte por la mitad.
    """
    suma = sum(ANCHOS[_letra(i)] for i in range(1, ULTIMA_COLUMNA_CARA_A + 1))
    if abs(suma - SUMA_ANCHOS_CARA_A) > 1e-9:
        raise ValueError(
            f"La suma de anchos de A a R es {suma}, debería ser "
            f"{SUMA_ANCHOS_CARA_A}. Si ensanchas una columna, quítale lo mismo "
            f"a otra.")
    for col, ancho in ANCHOS.items():
        if abs(ancho * 256 - round(ancho * 256)) > 1e-9:
            raise ValueError(
                f"El ancho de la columna {col} ({ancho}) no es representable: "
                f"Excel guarda los anchos en 1/256 de carácter.")
    return suma


def numero_de_lote(lote) -> int:
    """
    Número inicial de un lote, para ordenar. Los lotes son cadenas y muchos
    traen sufijo ('21107-A', '21113-1'): ordenados como texto, '21107-A'
    quedaría antes que '2200'. Los que no traen número se van al final.
    """
    m = re.match(r'\s*(\d+)', str(lote or ''))
    return int(m.group(1)) if m else 10 ** 9


def _crear_estilos():
    """
    Crea una sola vez cada juego de estilos. Con 40 filas x 38 columnas x N
    hojas, un Font o un Border nuevo por celda infla el archivo y hace lenta la
    generación.
    """
    from openpyxl.styles import Alignment, Border, Font, Side

    media = Side(style='medium')
    fina = Side(style='thin')

    e = {
        'titulo': Font(name='Calibri', sz=10, bold=True),
        'cab_8': Font(name='Arial', sz=8, bold=True),
        'cab_9': Font(name='Arial', sz=9, bold=True),
        'arial8': Font(name='Arial', sz=8),
        'arial9': Font(name='Arial', sz=9),
        'calibri9': Font(name='Calibri', sz=9),
        'centro': Alignment(horizontal='center'),
        'centro_medio': Alignment(horizontal='center', vertical='center'),
        # Cabeceras: recuadro completo. La mitad inferior de las columnas
        # combinadas no lleva línea arriba.
        'bd_cab': Border(left=media, right=media, top=media, bottom=media),
        'bd_cab_inf': Border(left=media, right=media, bottom=media),
    }
    # Bordes de los renglones de datos: laterales medios siempre; arriba medio
    # en el primer renglón de cada persona, abajo medio en el último. Es lo que
    # dibuja la rejilla de bloques de 4.
    e['bd_datos'] = [
        Border(left=media, right=media, top=media, bottom=fina),   # 1er renglón
        Border(left=media, right=media, top=fina, bottom=fina),
        Border(left=media, right=media, top=fina, bottom=fina),
        Border(left=media, right=media, top=fina, bottom=media),   # último
    ]
    return e


def _formato_columna_datos(idx: int, e: Dict):
    """Fuente, alineación y formato de número de una columna en los datos."""
    if idx == 2:                       # B - LOTE
        return e['calibri9'], e['centro'], 'General'
    if idx == 3:                       # C - NOMBRE
        return e['calibri9'], None, 'General'
    if idx == 6:                       # F - SUP.
        return e['calibri9'], e['centro'], '0.00'
    if idx == 7:                       # G - CULTIVO
        return e['arial8'], e['centro'], '@'
    if idx <= 16 or idx == COLUMNA_NUMERO_CARA_B:   # A, D..P y S
        return e['arial8'], e['centro'], 'General'
    return e['arial9'], None, 'General'              # Q, R y la zona de riegos


def _escribir_bloque(ws, base: int, personas: List[Dict], estilos: Dict,
                     logo_ruta: Optional[str]):
    """Escribe una página completa: cabeceras, logo y 40 renglones."""
    from openpyxl.drawing.image import Image as ImagenExcel

    e = estilos

    # --- Alturas de fila. Deben quedar fijas: si Excel las recalcula, el
    # bloque deja de medir 46 filas y la paginación se descuadra.
    for off in range(FILAS_POR_BLOQUE):
        ws.row_dimensions[base + off].height = ALTURAS.get(off, ALTURA_DATOS)

    # --- Título y subtítulo
    celda = ws.cell(row=base, column=6, value=TITULO)      # F
    celda.font = e['titulo']
    celda.alignment = e['centro']
    ws.merge_cells(start_row=base, start_column=6, end_row=base, end_column=15)

    celda = ws.cell(row=base + 2, column=6, value=SUBTITULO)
    celda.font = e['titulo']

    # --- Cabeceras (+4 y +5)
    fila_sup, fila_inf = base + 4, base + 5
    for col, texto in CABECERAS_SUP.items():
        celda = ws[f'{col}{fila_sup}']
        celda.value = texto
        celda.font = e['cab_8'] if celda.column <= 16 else e['cab_9']
        celda.alignment = e['centro_medio']
        celda.border = e['bd_cab']
    for col, texto in CABECERAS_INF.items():
        celda = ws[f'{col}{fila_inf}']
        celda.value = texto
        celda.font = e['cab_8'] if celda.column <= 16 else e['cab_9']
        celda.alignment = e['centro_medio']
        celda.border = e['bd_cab']
    # Mitad inferior de las columnas combinadas: sólo laterales y base.
    for col in COLUMNAS_MERGE_VERTICAL:
        ws[f'{col}{fila_inf}'].border = e['bd_cab_inf']

    # --- Combinaciones de las cabeceras
    for col in COLUMNAS_MERGE_VERTICAL:
        ws.merge_cells(f'{col}{fila_sup}:{col}{fila_inf}')
    ws.merge_cells(f'I{fila_sup}:K{fila_sup}')     # USO DE SEMILLA
    ws.merge_cells(f'M{fila_sup}:P{fila_sup}')     # USO DE MAQUINARIA
    ws.merge_cells(f'Q{fila_sup}:R{fila_sup}')     # 1ER. RIEGO
    for izq in range(PRIMER_RIEGO_CARA_B, PRIMER_RIEGO_CARA_B + 18, 2):  # 2º a 10º
        ws.merge_cells(start_row=fila_sup, start_column=izq,
                       end_row=fila_sup, end_column=izq + 1)

    # --- Renglones de datos: cuatro por persona, siempre.
    for i in range(FILAS_DATOS):
        fila = base + PRIMERA_FILA_DATOS + i
        posicion = i % RENGLONES_POR_PERSONA
        borde = e['bd_datos'][posicion]
        persona = personas[i // RENGLONES_POR_PERSONA] \
            if i // RENGLONES_POR_PERSONA < len(personas) else None

        for idx in range(1, TOTAL_COLUMNAS + 1):
            celda = ws.cell(row=fila, column=idx)
            fuente, alineacion, formato = _formato_columna_datos(idx, e)
            celda.font = fuente
            if alineacion is not None:
                celda.alignment = alineacion
            if formato != 'General':
                celda.number_format = formato
            celda.border = borde

        ws.cell(row=fila, column=1).value = i + 1          # A - consecutivo
        # El mismo consecutivo en la cara B: impreso a doble cara, frente y
        # reverso son las dos mitades de la misma página.
        ws.cell(row=fila, column=COLUMNA_NUMERO_CARA_B).value = i + 1

        # Sólo el primer renglón de cada persona lleva datos. Todo lo demás
        # (TEN., CULTIVO, semilla, maquinaria y los riegos) va vacío.
        if persona is not None and posicion == 0:
            ws.cell(row=fila, column=2).value = persona.get('numero_lote')
            ws.cell(row=fila, column=3).value = persona.get('nombre')
            ws.cell(row=fila, column=4).value = persona.get('barrio')
            superficie = persona.get('superficie')
            if superficie is not None:
                ws.cell(row=fila, column=6).value = float(superficie)

    # --- Logo, anclado en la A de la primera fila del bloque. Sólo en la cara A.
    if logo_ruta:
        imagen = ImagenExcel(logo_ruta)
        imagen.width = LOGO_ANCHO_PX
        imagen.height = LOGO_ALTO_PX
        ws.add_image(imagen, f'A{base}')


def _deduplicar_imagenes(ruta: str) -> int:
    """
    openpyxl incrusta los bytes del logo una vez por hoja: con 123 hojas el
    archivo se dispara. Deja una sola copia en xl/media y reapunta a ella todas
    las relaciones, respetando el estilo de ruta que ya traiga el archivo.

    Devuelve el número de imágenes que quedaron.
    """
    with zipfile.ZipFile(ruta) as z:
        entradas = [(info, z.read(info.filename)) for info in z.infolist()]

    medios = [(info.filename, datos) for info, datos in entradas
              if info.filename.startswith('xl/media/')]
    if len(medios) <= 1:
        return len(medios)

    canonico = {}          # bytes -> nombre que se conserva
    equivalencia = {}      # nombre base duplicado -> nombre base conservado
    for nombre, datos in medios:
        clave = (len(datos), datos)
        if clave in canonico:
            equivalencia[os.path.basename(nombre)] = os.path.basename(canonico[clave])
        else:
            canonico[clave] = nombre
    if not equivalencia:
        return len(medios)

    conservados = {nombre for nombre in canonico.values()}
    descartados = {nombre for nombre, _ in medios if nombre not in conservados}

    def reapuntar(xml: bytes) -> bytes:
        def sustituir(m):
            destino = m.group(1)
            base = destino.rsplit('/', 1)[-1]
            if base in equivalencia:
                destino = destino[:len(destino) - len(base)] + equivalencia[base]
            return f'Target="{destino}"'
        return re.sub(r'Target="([^"]*)"',
                      sustituir, xml.decode('utf-8')).encode('utf-8')

    temporal = ruta + '.tmp'
    with zipfile.ZipFile(temporal, 'w', zipfile.ZIP_DEFLATED) as salida:
        for info, datos in entradas:
            if info.filename in descartados:
                continue
            if info.filename.endswith('.rels'):
                datos = reapuntar(datos)
            salida.writestr(info, datos)
    shutil.move(temporal, ruta)
    return len(conservados)


def generar_libreta_excel(campesinos: List[Dict],
                          filename: Optional[str] = None) -> str:
    """
    Genera la libreta oficial en blanco con el padrón ya transcrito.

    Args:
        campesinos: lista de dicts con 'numero_lote', 'nombre', 'barrio' y
                    'superficie'. Se ordenan por número de lote ascendente.
        filename: nombre del archivo. Si no se da, se arma con la fecha.

    Returns:
        Ruta del archivo generado.
    """
    import openpyxl
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties

    verificar_anchos()

    if not campesinos:
        raise ValueError("No hay campesinos en el padrón para generar la libreta.")

    personas = sorted(campesinos,
                      key=lambda c: (numero_de_lote(c.get('numero_lote')),
                                     str(c.get('numero_lote') or '')))
    hojas = -(-len(personas) // PERSONAS_POR_HOJA)   # ceil

    logo_ruta = resource_path(LOGO_RUTA)
    if not os.path.exists(logo_ruta):
        logo_ruta = None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LIBRETA"

    for col, ancho in ANCHOS.items():
        ws.column_dimensions[col].width = ancho

    estilos = _crear_estilos()
    for hoja in range(hojas):
        base = hoja * FILAS_POR_BLOQUE + 1
        _escribir_bloque(
            ws, base,
            personas[hoja * PERSONAS_POR_HOJA:(hoja + 1) * PERSONAS_POR_HOJA],
            estilos, logo_ruta)

    # --- Impresión: copiado de 'cambios libreta.xlsx'. La escala del 95% no es
    # opcional: el bloque mide 578.5 pt y el área imprimible de un oficio
    # horizontal con estos márgenes son 563.8 pt.
    ws.page_setup.paperSize = 5                  # Legal (oficio)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.scale = 95
    ws.page_setup.pageOrder = 'overThenDown'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False)
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(
        left=0.5905511811023623, right=0.1968503937007874,
        top=0.3937007874015748, bottom=0.2755905511811024,
        header=0.31496062992125984, footer=0.31496062992125984)

    ws.col_breaks.append(Break(id=ULTIMA_COLUMNA_CARA_A))   # tras la R
    for n in range(1, hojas):
        ws.row_breaks.append(Break(id=FILAS_POR_BLOQUE * n))

    if filename is None:
        from datetime import datetime
        filename = f"libreta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    filepath = os.path.join(reportes_dir, filename)
    wb.save(filepath)

    if logo_ruta:
        _deduplicar_imagenes(filepath)

    return filepath
